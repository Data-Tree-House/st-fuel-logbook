"""Utilities for generating a personalised bulk-upload template workbook."""

import shutil
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from db.model import Car

# ── Sheet names ────────────────────────────────────────────────────────────────
TEMPLATE_INPUT_SHEET = "Input"
FUEL_ENTRIES_SHEET = "Fuel Entries"
VEHICLES_SHEET = "Vehicles"
METADATA_SHEET = "metadata"

# ── Vehicles sheet: ordered column headers (1-based positions) ────────────────
VEHICLES_HEADERS: list[str] = [
    "Id",
    "Nickname",
    "Fuel Type",
    "Registration Number",
    "VIN Number",
    "Model Description",
    "Color",
    "Registration Date",
]

# ── Column letter for Vehicle in the Fuel Entries sheet ───────────────────────
FUEL_ENTRIES_VEHICLE_COL_LETTER = "B"
# Header of the Fuel Type column to remove from Fuel Entries
FUEL_ENTRIES_FUEL_TYPE_HEADER = "Fuel Type"

MAX_DATA_ROWS = 10_000


def populate_template(template_path: Path, cars: list[Car]) -> bytes:
    """Generate a personalised copy of the upload template for the given user.

    Steps performed:
    - Renames the "Input" sheet to "Fuel Entries" (idempotent).
    - Adds a "Fuel Type" header to the Vehicles sheet if not already present.
    - Populates the "Vehicles" sheet rows with the user's cars (Id, Nickname, Fuel Type).
    - Adds data-validation dropdowns:
        • Vehicle column in "Fuel Entries"  →  Vehicles!$B$2:$B$N
        • Fuel Type column in "Vehicles"    →  metadata!$A$1:$A$5

    Args:
        template_path: Path to the base template ``.xlsx`` file.
        cars: The logged-in user's (non-deleted) cars.

    Returns:
        The personalised workbook as raw bytes ready to be served for download.
    """
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = Path(tmp.name)

    shutil.copy2(template_path, tmp_path)
    try:
        wb = openpyxl.load_workbook(tmp_path)

        _rename_input_sheet(wb)
        _remove_fuel_type_from_fuel_entries(wb)
        _populate_vehicles_sheet(wb, cars)
        _add_fuel_type_dropdown_to_vehicles(wb)
        _add_vehicle_dropdown_to_fuel_entries(wb, cars)

        wb.save(tmp_path)
        return tmp_path.read_bytes()
    finally:
        tmp_path.unlink(missing_ok=True)


# ── Private helpers ────────────────────────────────────────────────────────────


def _rename_input_sheet(wb: openpyxl.Workbook) -> None:
    """Rename the legacy "Input" sheet to "Fuel Entries" if needed."""
    if TEMPLATE_INPUT_SHEET in wb.sheetnames and FUEL_ENTRIES_SHEET not in wb.sheetnames:
        wb[TEMPLATE_INPUT_SHEET].title = FUEL_ENTRIES_SHEET


def _remove_fuel_type_from_fuel_entries(wb: openpyxl.Workbook) -> None:
    """Delete the "Fuel Type" column from the Fuel Entries sheet if present.

    Fuel type belongs to the vehicle, not to individual log entries.
    """
    if FUEL_ENTRIES_SHEET not in wb.sheetnames:
        return

    ws = wb[FUEL_ENTRIES_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    for col_idx, header in enumerate(header_row, start=1):
        if header == FUEL_ENTRIES_FUEL_TYPE_HEADER:
            ws.delete_cols(col_idx)
            break


def _populate_vehicles_sheet(wb: openpyxl.Workbook, cars: list[Car]) -> None:
    """Write the user's cars into the Vehicles sheet.

    Writes all ``VEHICLES_HEADERS`` columns and populates one row per car.
    Existing data rows are cleared before writing.
    """
    ws = wb[VEHICLES_SHEET]

    # Write / overwrite the full header row
    for col_idx, header in enumerate(VEHICLES_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)

    # Clear any stale data rows
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.value = None

    # Map header name → car attribute for concise population
    car_field_map: dict[str, str] = {
        "Id": "id",
        "Nickname": "nickname",
        "Fuel Type": "fuel_type",
        "Registration Number": "registration_number",
        "VIN Number": "vin_number",
        "Model Description": "model_description",
        "Color": "color",
        "Registration Date": "registration_date",
    }

    for row_idx, car in enumerate(cars, start=2):
        for col_idx, header in enumerate(VEHICLES_HEADERS, start=1):
            attr = car_field_map[header]
            ws.cell(row=row_idx, column=col_idx, value=getattr(car, attr))


def _add_fuel_type_dropdown_to_vehicles(wb: openpyxl.Workbook) -> None:
    """Add a dropdown on the Fuel Type column of the Vehicles sheet.

    The dropdown source comes from the existing ``metadata`` sheet which
    lists all valid fuel types.
    """
    if METADATA_SHEET not in wb.sheetnames:
        return

    ws = wb[VEHICLES_SHEET]
    fuel_type_col_idx = VEHICLES_HEADERS.index("Fuel Type") + 1  # 1-based
    fuel_type_col_letter = get_column_letter(fuel_type_col_idx)

    # Remove any existing validation on this column
    ws.data_validations.dataValidation = [
        dv for dv in ws.data_validations.dataValidation if fuel_type_col_letter not in str(dv.sqref)
    ]

    # Count how many fuel types exist in metadata
    meta_ws = wb[METADATA_SHEET]
    fuel_type_count = sum(1 for row in meta_ws.iter_rows(values_only=True) if row[0] is not None)
    last_meta_row = max(fuel_type_count, 1)

    dv = DataValidation(
        type="list",
        formula1=f"{METADATA_SHEET}!$A$1:$A${last_meta_row}",
        allow_blank=True,
        showErrorMessage=True,
        error="Please select a valid fuel type from the list.",
        errorTitle="Invalid Fuel Type",
    )
    dv.sqref = f"{fuel_type_col_letter}2:{fuel_type_col_letter}{MAX_DATA_ROWS}"
    ws.add_data_validation(dv)


def _add_vehicle_dropdown_to_fuel_entries(wb: openpyxl.Workbook, cars: list[Car]) -> None:
    """Add a dropdown on the Vehicle column of the Fuel Entries sheet.

    The dropdown source is the Nickname column of the Vehicles sheet.
    """
    ws = wb[FUEL_ENTRIES_SHEET]
    col = FUEL_ENTRIES_VEHICLE_COL_LETTER

    # Remove any existing validation on the Vehicle column
    ws.data_validations.dataValidation = [dv for dv in ws.data_validations.dataValidation if col not in str(dv.sqref)]

    last_vehicle_row = max(len(cars) + 1, 2)

    dv = DataValidation(
        type="list",
        formula1=f"{VEHICLES_SHEET}!$B$2:$B${last_vehicle_row}",
        allow_blank=True,
        showErrorMessage=True,
        error="Please select a valid vehicle from the list.",
        errorTitle="Invalid Vehicle",
    )
    dv.sqref = f"{col}2:{col}{MAX_DATA_ROWS}"
    ws.add_data_validation(dv)
