"""Utilities for parsing a bulk-upload workbook into ORM-ready objects."""

from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook

from db.model import FuelEntry

# ── Sheet / column constants ───────────────────────────────────────────────────
FUEL_ENTRIES_SHEET = "Fuel Entries"
VEHICLES_SHEET = "Vehicles"

REQUIRED_SHEETS: frozenset[str] = frozenset({FUEL_ENTRIES_SHEET, VEHICLES_SHEET})

# All columns expected in the Fuel Entries sheet
# Note: Fuel Type lives on the vehicle (Vehicles sheet), not on individual log entries.
FUEL_ENTRIES_COLUMNS: list[str] = [
    "Date (DD/MM/YYYY)",
    "Vehicle",
    "Odometer (km)",
    "Trip Distance (km)",
    "Fuel Filled (Liters)",
    "Price",
    "Location",
]


# ── Data classes ───────────────────────────────────────────────────────────────


@dataclass
class ParsedVehicle:
    """A vehicle row parsed from the Vehicles sheet.

    Attributes:
        nickname: The display name of the vehicle.
        id: The database ``Car.id``.  ``None`` means the car does not yet exist
            in the database and must be created during the upload.
        fuel_type: Required when ``id`` is ``None`` (new vehicle).
        registration_number: Optional vehicle registration plate.
        vin_number: Optional 17-character VIN.
        model_description: Optional free-text model description.
        color: Optional color string.
        registration_date: Optional date the vehicle was first registered.
    """

    nickname: str
    id: int | None = None
    fuel_type: str | None = None
    registration_number: str | None = None
    vin_number: str | None = None
    model_description: str | None = None
    color: str | None = None
    registration_date: date | None = None


@dataclass
class ParseError:
    """A single row-level parsing or validation error.

    Attributes:
        row: The 1-based Excel row number where the error occurred.
        error: A human-readable description of the error.
    """

    row: int
    error: str


# ── Public helpers ─────────────────────────────────────────────────────────────


def load_workbook_from_bytes(data: bytes) -> Workbook:
    """Deserialise raw bytes into an :class:`openpyxl.Workbook`.

    Args:
        data: Raw ``.xlsx`` file bytes.

    Returns:
        An openpyxl Workbook instance.
    """
    return openpyxl.load_workbook(BytesIO(data))


def validate_workbook(wb: Workbook) -> tuple[bool, str]:
    """Check that the workbook has the expected sheets and Fuel Entries columns.

    Args:
        wb: The uploaded workbook.

    Returns:
        A ``(is_valid, error_message)`` tuple.  ``error_message`` is an empty
        string when ``is_valid`` is ``True``.
    """
    missing_sheets = REQUIRED_SHEETS - set(wb.sheetnames)
    if missing_sheets:
        return (
            False,
            f"Missing required sheets: {', '.join(sorted(missing_sheets))}. Please use the provided template.",
        )

    fuel_ws = wb[FUEL_ENTRIES_SHEET]
    header_row: tuple[Any, ...] = next(fuel_ws.iter_rows(min_row=1, max_row=1, values_only=True))  # type: ignore[assignment]
    existing_headers = {str(cell) for cell in header_row if cell is not None}
    missing_cols = set(FUEL_ENTRIES_COLUMNS) - existing_headers
    if missing_cols:
        return (
            False,
            f"Missing columns in '{FUEL_ENTRIES_SHEET}': {', '.join(sorted(missing_cols))}.",
        )

    has_data = any(any(cell is not None for cell in row) for row in fuel_ws.iter_rows(min_row=2, values_only=True))
    if not has_data:
        return False, f"The '{FUEL_ENTRIES_SHEET}' sheet contains no data rows."

    return True, ""


def parse_vehicles_sheet(wb: Workbook) -> tuple[list[ParsedVehicle], list[ParseError]]:
    """Parse the Vehicles sheet into :class:`ParsedVehicle` objects.

    Column positions are resolved from the header row, so the sheet is
    compatible with both the legacy 2-column format and the full 8-column
    template layout.

    Rows with an empty ``Nickname`` cell are silently skipped.
    Rows with an ``Id`` of ``None`` represent new vehicles to be created.

    Args:
        wb: The uploaded workbook.

    Returns:
        A tuple of ``(vehicles, errors)``.
    """
    ws = wb[VEHICLES_SHEET]

    # Build a header → column-index map from row 1
    header_row: tuple[Any, ...] = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # type: ignore[assignment]
    col: dict[str, int] = {str(h).strip(): i for i, h in enumerate(header_row) if h is not None}

    def _get(row: tuple[Any, ...], header: str) -> Any:
        """Return the cell value for *header*, or ``None`` if the column is absent."""
        idx = col.get(header)
        return row[idx] if idx is not None and idx < len(row) else None

    def _str_or_none(value: Any) -> str | None:
        if value is None:
            return None
        stripped = str(value).strip()
        return stripped if stripped else None

    def _date_or_none(value: Any) -> date | None:
        if value is None:
            return None
        if isinstance(value, date):
            return value
        try:
            from datetime import datetime

            return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()  # noqa: DTZ007
        except (ValueError, TypeError):
            return None

    vehicles: list[ParsedVehicle] = []
    errors: list[ParseError] = []

    for excel_row, row_raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row: tuple[Any, ...] = row_raw  # type: ignore[assignment]

        nickname_raw = _get(row, "Nickname")
        if nickname_raw is None:
            continue  # Blank row - skip silently

        nickname = str(nickname_raw).strip()
        if not nickname:
            continue

        # Parse the optional id
        vehicle_id_raw = _get(row, "Id")
        parsed_id: int | None
        if vehicle_id_raw is None:
            parsed_id = None
        else:
            try:
                parsed_id = int(vehicle_id_raw)
            except (ValueError, TypeError):
                errors.append(
                    ParseError(
                        row=excel_row,
                        error=f"Invalid Id value '{vehicle_id_raw}' for vehicle '{nickname}'.",
                    )
                )
                continue

        # New vehicles must have a fuel type
        fuel_type = _str_or_none(_get(row, "Fuel Type"))
        if parsed_id is None and not fuel_type:
            errors.append(
                ParseError(
                    row=excel_row,
                    error=f"New vehicle '{nickname}' is missing a Fuel Type.",
                )
            )
            continue

        vehicles.append(
            ParsedVehicle(
                nickname=nickname,
                id=parsed_id,
                fuel_type=fuel_type,
                registration_number=_str_or_none(_get(row, "Registration Number")),
                vin_number=_str_or_none(_get(row, "VIN Number")),
                model_description=_str_or_none(_get(row, "Model Description")),
                color=_str_or_none(_get(row, "Color")),
                registration_date=_date_or_none(_get(row, "Registration Date")),
            )
        )

    return vehicles, errors


def validate_vehicle_references(
    wb: Workbook,
    known_nicknames: set[str],
) -> list[ParseError]:
    """Ensure every nickname used in the Fuel Entries sheet exists in *known_nicknames*.

    Args:
        wb: The uploaded workbook.
        known_nicknames: The set of all valid nicknames (from the Vehicles sheet).

    Returns:
        A (possibly empty) list of :class:`ParseError` for unresolvable rows.
    """
    ws = wb[FUEL_ENTRIES_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    vehicle_col_idx = headers.index("Vehicle")

    errors: list[ParseError] = []
    for excel_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue

        nickname_raw = row[vehicle_col_idx]
        if nickname_raw is None or str(nickname_raw).strip() == "":
            errors.append(ParseError(row=excel_row, error="Vehicle is required."))
        elif str(nickname_raw).strip() not in known_nicknames:
            errors.append(
                ParseError(
                    row=excel_row,
                    error=f"Vehicle '{nickname_raw}' not found in the Vehicles sheet.",
                )
            )

    return errors


def parse_fuel_entries_sheet(
    wb: Workbook,
    nickname_to_car_id: dict[str, int],
) -> tuple[list[FuelEntry], list[ParseError]]:
    """Parse the Fuel Entries sheet into :class:`~db.model.FuelEntry` ORM objects.

    Note: The ``Fuel Type`` column is intentionally ignored here because fuel
    type is stored on the :class:`~db.model.Car`, not on individual entries.

    Args:
        wb: The uploaded workbook.
        nickname_to_car_id: A mapping of vehicle nickname → ``Car.id``.
            This must include both pre-existing and newly flushed cars.

    Returns:
        A tuple of ``(fuel_entries, errors)``.
    """
    ws = wb[FUEL_ENTRIES_SHEET]
    header_row: tuple[Any, ...] = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))  # type: ignore[assignment]
    col: dict[str, int] = {str(h): i for i, h in enumerate(header_row) if h is not None}

    entries: list[FuelEntry] = []
    errors: list[ParseError] = []

    for excel_row, row_raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row: tuple[Any, ...] = row_raw  # type: ignore[assignment]
        if all(cell is None for cell in row):
            continue

        try:
            nickname = str(row[col["Vehicle"]]).strip()
            car_id = nickname_to_car_id[nickname]

            entry = FuelEntry(
                car_id=car_id,
                entry_datetime=row[col["Date (DD/MM/YYYY)"]],
                odometer=float(row[col["Odometer (km)"]]),
                trip=float(row[col["Trip Distance (km)"]]),
                fuel_filled=float(row[col["Fuel Filled (Liters)"]]),
                price=float(row[col["Price"]]),
                location=row[col["Location"]],
            )
            entries.append(entry)

        except KeyError as e:
            errors.append(ParseError(row=excel_row, error=f"Unknown vehicle: {e}"))
        except (ValueError, TypeError) as e:
            errors.append(ParseError(row=excel_row, error=str(e)))
        except Exception as e:
            errors.append(ParseError(row=excel_row, error=f"Unexpected error: {e!s}"))

    return entries, errors
